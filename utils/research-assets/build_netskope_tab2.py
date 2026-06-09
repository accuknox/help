import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from collections import Counter

wb = openpyxl.load_workbook(r'C:\Users\AtharvaShah\Downloads\AI Security_ AccuKnox vs NetSkope.xlsx')

# Remove old tab if it exists
if 'Detailed Comparison' in wb.sheetnames:
    del wb['Detailed Comparison']

ws = wb.create_sheet("Detailed Comparison")

hdr_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
hdr_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
cat_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
cell_font = Font(name='Calibri', size=10)
supp_link_font = Font(name='Calibri', size=10, color='0563C1', underline='single')
beta_link_font = Font(name='Calibri', size=10, italic=True, color='C27C1B', underline='single')
wrap_top = Alignment(wrap_text=True, vertical='top')
wrap_mid = Alignment(wrap_text=True, vertical='center', horizontal='center')
border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
thick_border_bottom = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='medium')
)
win_fills = {
    'AccuKnox': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
    'Netskope': PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
    'Split':    PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'),
    'Parity':   PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'),
}

# Headers — no separate Docs column; link lives in column C
headers = ['Category', 'Feature', 'AccuKnox', 'Netskope', 'Winner', 'AccuKnox Differentiation']
for col, h in enumerate(headers, 1):
    c = ws.cell(row=1, column=col, value=h)
    c.font = hdr_font
    c.fill = hdr_fill
    c.alignment = wrap_top
    c.border = border

ws.column_dimensions['A'].width = 28
ws.column_dimensions['B'].width = 40
ws.column_dimensions['C'].width = 16
ws.column_dimensions['D'].width = 16
ws.column_dimensions['E'].width = 14
ws.column_dimensions['F'].width = 70
ws.row_dimensions[1].height = 22

# (category, feature, accuknox_status, accuknox_link, netskope, winner, differentiation)
data = [
    # 1. Governance & AI-SPM
    ('Governance & AI-SPM', 'AI Asset Discovery & Inventory',
     'Supported', 'https://help.accuknox.com/how-to/aiml-overview/',
     'Supported', 'AccuKnox',
     'Security Graph pipeline view. Agents mapped to linked models, knowledge bases, tools. Discovers on-prem inference engines and their hosted models. Covers LLMs and ML models.'),

    ('Governance & AI-SPM', 'Secure Use of GenAI Apps (ChatGPT, Copilot)',
     'Supported', 'https://help.accuknox.com/integrations/openai-browser-integration/',
     'Supported', 'Parity',
     'Browser plugin (Chrome + Firefox), CLI, API, AI Gateway integration modes.'),

    ('Governance & AI-SPM', 'Unauthorized SaaS AI App Control',
     'Supported', 'https://help.accuknox.com/how-to/aiml-overview/',
     'Supported', 'Netskope',
     'Netskope has deeper CASB heritage for SaaS app discovery and control. AccuKnox covers this via AI-SPM.'),

    ('Governance & AI-SPM', 'Shadow AI Detection',
     'Supported', 'https://help.accuknox.com/how-to/aiml-overview/',
     'Supported', 'Split',
     'AccuKnox detects self-hosted inference engines (Ollama, vLLM) and local AI deployments with workflow alerts. Netskope tracks SaaS AI app usage via CASB.'),

    # 2. Model Security
    ('Model Security', 'AI Red Teaming',
     'Supported', 'https://help.accuknox.com/use-cases/red-teaming/',
     'Not Supported', 'AccuKnox',
     'Upload custom domain-specific adversarial prompts. Automated red-teaming with configurable attack categories.'),

    ('Model Security', 'LLM Static Scanning',
     'Supported', 'https://help.accuknox.com/how-to/llm-static-scan/',
     'Not Supported', 'AccuKnox',
     'Evaluate models on: provenance, adversarial robustness, supply chain, AIBOM, model file security (pickle, serialization/deserialization).'),

    ('Model Security', 'ML Model Static Scanning',
     'Supported', 'https://help.accuknox.com/how-to/ml-static-scan/',
     'Not Supported', 'AccuKnox',
     'Static security analysis for traditional ML models beyond LLMs.'),

    # 3. AI Supply Chain
    ('AI Supply Chain', 'AIBOM / xBOM',
     'Supported', 'https://help.accuknox.com/getting-started/xbom-setup/',
     'Not Supported', 'AccuKnox',
     'SBOM, CBOM, AIBOM, HBOM: generate, create, compare, sign, verify. Supports CERT-In guidelines.'),

    ('AI Supply Chain', 'Model Poisoning Detection',
     'Supported', 'https://help.accuknox.com/use-cases/modelarmor/',
     'Not Supported', 'AccuKnox',
     'Detects poisoning in locally deployed inference engines (Ollama, vLLM). ModelArmor open-source tool.'),

    ('AI Supply Chain', 'AI Pipeline Security',
     'Supported', 'https://help.accuknox.com/how-to/aiml-overview/',
     'Not Supported', 'AccuKnox',
     'Full pipeline visibility: training data, model training, serving infrastructure. Security Graph maps dependencies across the AI stack.'),

    # 4. Runtime Security & Guardrails
    ('Runtime Security & Guardrails', 'Prompt Firewalling',
     'Supported', 'https://help.accuknox.com/use-cases/prompt-firewall-overview/',
     'Supported', 'AccuKnox',
     'ModelArmor OSS + enterprise. Semantic detection across custom sub-prompt categories. SDK, API, and Gateway modes. Netskope uses DLP-based inline inspection only.'),

    ('Runtime Security & Guardrails', 'Safety Guardrails: Session Abuse',
     'Supported', 'https://help.accuknox.com/use-cases/subprompts-categories/',
     'Partial', 'AccuKnox',
     'Session-level abuse detection, jailbreak pattern recognition, multi-turn attack detection. Netskope limited to basic DLP pattern matching.'),

    ('Runtime Security & Guardrails', 'Safety Guardrails: Unsafe Content',
     'Supported', 'https://help.accuknox.com/use-cases/prompt-firewall-overview/',
     'Partial', 'AccuKnox',
     'Output filtering, toxicity detection, custom content categories. Netskope limited to DLP-based output inspection.'),

    ('Runtime Security & Guardrails', 'AI Agent Sandboxing [Beta]',
     'Beta', 'https://accuknox.com/ai-security/',
     'Not Supported', 'AccuKnox',
     'OpenClaw: autonomous agent sandboxing using KubeArmor-based eBPF enforcement. Isolates agent tool calls and file system access.'),

    ('Runtime Security & Guardrails', 'MCP Server Execution Sandboxing [Beta]',
     'Beta', 'https://help.accuknox.com/integrations/mcp-server/',
     'Not Supported', 'AccuKnox',
     'Sandbox MCP server tool executions. Prevent unauthorized system calls from MCP-connected tools.'),

    ('Runtime Security & Guardrails', 'Untrusted Model Sandboxing [Beta]',
     'Beta', 'https://accuknox.com/ai-security/',
     'Not Supported', 'AccuKnox',
     'Isolate untrusted model inference in sandboxed environments with restricted system access.'),

    # 5. AI Data Security
    ('AI Data Security', 'Data Poisoning / RAG Fencing',
     'Supported', 'https://help.accuknox.com/use-cases/aiml-usecases/',
     'Not Supported', 'AccuKnox',
     'Protect training data from poisoning. Protect RAG datasets from poisoning. Fence sensitive data from RAG retrieval.'),

    ('AI Data Security', 'AI Sovereignty / Regional Compliance',
     'Supported', 'https://help.accuknox.com/use-cases/aidr/',
     'Supported', 'Parity',
     'AI asset threat detection rules for regional settings. Netskope supports data residency via CASB controls.'),

    # 6. AI Detection & Response
    ('AI Detection & Response', 'Cloud Log AI Threat Detection (AWS)',
     'Supported', 'https://help.accuknox.com/use-cases/aidr/',
     'Not Supported', 'AccuKnox',
     'Consume AWS cloud logs, detect threats on AI assets (public exposure, unauthorized access, config drift), auto-remediate.'),

    ('AI Detection & Response', 'Cloud Log AI Threat Detection (Azure)',
     'Supported', 'https://help.accuknox.com/use-cases/azure-aidr/',
     'Not Supported', 'AccuKnox',
     'Azure-specific AI asset threat detection rules and automated response.'),

    ('AI Detection & Response', 'Incident Response / Auto-Remediation',
     'Supported', 'https://help.accuknox.com/use-cases/aidr/',
     'Partial', 'AccuKnox',
     'Auto-remediate: change exposed assets to private, alert data owners/custodians/security teams, create ITSM tickets (Jira, ServiceNow). Netskope limited to alert-based notifications.'),

    # 7. AI Gateway & Integrations
    ('AI Gateway & Integrations', 'AI Gateway Integrations',
     'Supported', 'https://help.accuknox.com/integrations/ai-overview/',
     'Partial', 'AccuKnox',
     'Azure AI Foundry, AWS API Management, Apigee, Bifrost, LiteLLM. Netskope uses inline proxy with no dedicated AI gateway connectors.'),

    ('AI Gateway & Integrations', 'Copilot Studio / Power Apps',
     'Supported', 'https://help.accuknox.com/integrations/copilot-studio/',
     'Not Supported', 'AccuKnox',
     'Direct integration with Microsoft Copilot Studio and Power Apps for enterprise AI workflows.'),

    ('AI Gateway & Integrations', 'AWS Bedrock AgentCore',
     'Supported', 'https://help.accuknox.com/integrations/bedrock-agentcore/',
     'Not Supported', 'AccuKnox',
     'Native integration with AWS Bedrock AgentCore for securing AWS-hosted AI agents.'),

    ('AI Gateway & Integrations', 'SDK and Platform Integrations',
     'Supported', 'https://help.accuknox.com/integrations/ai-overview/',
     'Supported', 'AccuKnox',
     'SDK, Browser Plugin (Chrome + Firefox), CLI, API, Gateway proxy. Wider AI-specific integration surface than Netskope CASB connectors.'),

    # 8. Securing AI Factories
    ('Securing AI Factories', 'Securing NVIDIA RunAI',
     'Supported', 'https://accuknox.com/ai-security/',
     'Not Supported', 'AccuKnox',
     'Dedicated AI Factory security solution for NVIDIA RunAI environments. Beta docs coming soon.'),

    ('Securing AI Factories', 'Multi-Tenant GPU Isolation',
     'Supported', 'https://accuknox.com/ai-security/',
     'Not Supported', 'AccuKnox',
     'Isolate GPU workloads across tenants in shared AI infrastructure. Beta docs coming soon.'),

    ('Securing AI Factories', 'Authorized Access to CUDA',
     'Supported', 'https://accuknox.com/ai-security/',
     'Not Supported', 'AccuKnox',
     'Control and audit access to CUDA runtime and GPU resources. Beta docs coming soon.'),

    # 9. Deployment & Operations
    ('Deployment & Operations', 'On-Prem / Air-Gapped Deployment',
     'Supported', 'https://help.accuknox.com/getting-started/on-prem-overview/',
     'Not Supported', 'AccuKnox',
     'Full on-prem, air-gapped, edge, single-node deployment options. Netskope AI security is SaaS-only.'),

    ('Deployment & Operations', 'Deployment Flexibility',
     'Supported', 'https://help.accuknox.com/getting-started/deployment-models/',
     'Partial', 'AccuKnox',
     'SaaS, on-prem, hybrid, MSSP multi-tenant. Netskope offers SaaS with NewEdge PoPs but no on-prem AI security option.'),

    # 10. Agentic AI Security (all Beta)
    ('Agentic AI Security [Beta]', 'Agent Artifact Scanning [Beta]',
     'Beta', 'https://accuknox.com/ai-security/',
     'Not Supported', 'AccuKnox',
     'Scan agent artifacts (tool definitions, schemas, configurations) for security risks before deployment.'),

    ('Agentic AI Security [Beta]', 'Agent Red Teaming [Beta]',
     'Beta', 'https://accuknox.com/ai-security/',
     'Not Supported', 'AccuKnox',
     'Adversarial testing of AI agent behaviors: tool misuse, privilege escalation, data exfiltration through agentic workflows.'),

    ('Agentic AI Security [Beta]', 'Agent Security Posture Management [Beta]',
     'Beta', 'https://accuknox.com/ai-security/',
     'Not Supported', 'AccuKnox',
     'Inventory and assess security posture of deployed AI agents, their tool access, and permission scopes.'),

    ('Agentic AI Security [Beta]', 'Agentic Runtime Security [Beta]',
     'Beta', 'https://accuknox.com/ai-security/',
     'Not Supported', 'AccuKnox',
     'Runtime monitoring and enforcement for autonomous AI agent actions. eBPF-based observability via KubeArmor.'),

    ('Agentic AI Security [Beta]', 'Agent Observability [Beta]',
     'Beta', 'https://accuknox.com/ai-security/',
     'Not Supported', 'AccuKnox',
     'Trace and log AI agent decision chains, tool invocations, and data flows for audit and forensics.'),

    ('Agentic AI Security [Beta]', 'Agentic Identity Security [Beta]',
     'Beta', 'https://accuknox.com/ai-security/',
     'Not Supported', 'AccuKnox',
     'Identity and access management for AI agents. Control which agents can access which tools and data sources.'),

    ('Agentic AI Security [Beta]', 'Agentic Endpoint Security [Beta]',
     'Beta', 'https://accuknox.com/ai-security/',
     'Not Supported', 'AccuKnox',
     'Protect endpoints where AI agents execute: local machines, CI/CD runners, cloud VMs.'),

    ('Agentic AI Security [Beta]', 'AI Agent Gateway [Beta]',
     'Beta', 'https://help.accuknox.com/integrations/ai-overview/',
     'Not Supported', 'AccuKnox',
     'Centralized gateway for AI agent traffic. Policy enforcement for agent-to-tool and agent-to-model communications. Portkey integration.'),
]

# Track category row ranges for merging later
cat_start_row = {}
prev_cat = None
row = 2

for item in data:
    cat, feat, ak_status, ak_link, ns, winner, diff = item

    if cat != prev_cat:
        cat_start_row[cat] = row
        prev_cat = cat

    # Column A — category (value only; styled & merged after)
    ws.cell(row=row, column=1, value=cat)

    # Column B — feature
    b = ws.cell(row=row, column=2, value=feat)
    b.font = cell_font

    # Column C — AccuKnox status with embedded hyperlink
    c = ws.cell(row=row, column=3, value=ak_status)
    c.hyperlink = ak_link
    c.font = beta_link_font if ak_status == 'Beta' else supp_link_font

    # Column D — Netskope
    d = ws.cell(row=row, column=4, value=ns)
    d.font = cell_font

    # Column E — Winner
    e = ws.cell(row=row, column=5, value=winner)
    e.font = Font(name='Calibri', size=10, bold=True)
    if winner in win_fills:
        e.fill = win_fills[winner]

    # Column F — Differentiation
    f = ws.cell(row=row, column=6, value=diff)
    f.font = cell_font

    for col in range(1, 7):
        ws.cell(row=row, column=col).alignment = wrap_top
        ws.cell(row=row, column=col).border = border

    row += 1

last_data_row = row - 1

# Merge column A cells per category and style the merged cell
cats_in_order = list(dict.fromkeys(item[0] for item in data))
for i, cat in enumerate(cats_in_order):
    start = cat_start_row[cat]
    if i + 1 < len(cats_in_order):
        end = cat_start_row[cats_in_order[i + 1]] - 1
    else:
        end = last_data_row

    if start < end:
        ws.merge_cells(start_row=start, start_column=1, end_row=end, end_column=1)

    mc = ws.cell(row=start, column=1)
    mc.value = cat
    mc.font = Font(name='Calibri', bold=True, size=10)
    mc.fill = cat_fill
    mc.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
    # Thick bottom border on last row of each category block
    for r in range(start, end + 1):
        for col in range(1, 7):
            cell = ws.cell(row=r, column=col)
            is_last = (r == end)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='medium' if is_last else 'thin')
            )

# Scorecard rows
row += 1
winners_count = Counter(item[5] for item in data)
summary = "AccuKnox leads: {}  |  Netskope leads: {}  |  Split: {}  |  Parity: {}  |  Total: {}".format(
    winners_count.get('AccuKnox', 0), winners_count.get('Netskope', 0),
    winners_count.get('Split', 0), winners_count.get('Parity', 0), len(data)
)
sc_label = ws.cell(row=row, column=1, value='SCORECARD')
sc_label.font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
sc_label.fill = hdr_fill
sc_label.border = border
sc_label.alignment = wrap_top

sc = ws.cell(row=row, column=2, value=summary)
sc.font = Font(name='Calibri', bold=True, size=10)
ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
sc.border = border
sc.alignment = wrap_top

row += 1
edge = (
    "AccuKnox edge: model/dataset/pipeline artifact security, on-prem deployment, eBPF runtime "
    "enforcement (KubeArmor), AI detection & response with auto-remediation, AI factory security, "
    "full agentic AI security roadmap.\n"
    "Netskope edge: CASB-based SaaS AI app governance, DLP for GenAI, Cloud Confidence Index for shadow AI."
)
ws.cell(row=row, column=1).border = border
ec = ws.cell(row=row, column=2, value=edge)
ec.font = cell_font
ec.alignment = Alignment(wrap_text=True, vertical='top')
ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
ws.row_dimensions[row].height = 48

ws.freeze_panes = 'A2'

wb.save(r'C:\Users\AtharvaShah\Downloads\AI Security_ AccuKnox vs NetSkope.xlsx')
print("Saved. {} features, {} categories.".format(len(data), len(cats_in_order)))
print(summary)
