"""
Add an Appendix tab "Appendix Additional DAST and API Security Features"
to the Vodafone DAST Technical Evaluation workbook.

Each row lists an AccuKnox DAST/API security capability that is NOT already
covered in the existing " DAST Tech Specifications" tab, with a direct link
into help.accuknox.com docs.
"""

import copy
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SRC = r"C:\Users\AtharvaShah\Downloads\DAST Technical Evalution - Vodafone India - 21 May_2026.xlsx"
DST = r"C:\Users\AtharvaShah\Downloads\DAST Technical Evalution - Vodafone India - 21 May_2026.xlsx"

BASE = "https://help.accuknox.com"

# (Category, Feature, Description, AccuKnox Capability Details, Doc path on help.accuknox.com)
ROWS = [
    # 1: Technical Integration
    ("1: Technical Integration",
     "CI/CD-Integrated DAST (10+ Platforms)",
     "Native DAST execution from pipelines including GitHub, GitLab, Jenkins, Azure DevOps, Bitbucket, CircleCI, Bamboo, Harness, AWS CodePipeline, and Google Cloud Build.",
     "AccuKnox ships first-party DAST steps for 10+ CI/CD platforms. Pipelines trigger scans, push findings into the AccuKnox SaaS console, and gate builds on severity thresholds. The same scanner image runs locally via knoxctl for parity between developer machines and CI.",
     "/integrations/cicd-overview/"),

    ("1: Technical Integration",
     "DevSecOps Shift-Left Workflow",
     "End-to-end DevSecOps coverage spanning SAST, DAST, SCA, IaC, container scan, and secrets, with code-to-runtime risk correlation.",
     "AccuKnox correlates DAST findings with SAST, SCA, IaC, container, and runtime signals so the same CVE is not triaged five times. Severity, EPSS, CISA KEV, and runtime exposure feed a single risk score.",
     "/getting-started/devsecops/"),

    ("1: Technical Integration",
     "SARIF Ingestion (Multi-Tool ASPM)",
     "Ingest SARIF results from any third-party DAST/SAST tool so findings live alongside AccuKnox-native scan results.",
     "AccuKnox accepts SARIF and Checkmarx One uploads via API or pipeline step. Findings normalize into the same data model used by AccuKnox DAST, so dashboards, rules, and ticketing apply uniformly.",
     "/getting-started/sarif-findings/"),

    ("1: Technical Integration",
     "knoxctl CLI for DAST/API Scans",
     "Single-binary CLI to trigger DAST scans, generate xBOMs, and push findings without using the web UI.",
     "knoxctl runs DAST and container/SCA scans locally or in any pipeline runner, then forwards results to the AccuKnox control plane. Suited to air-gapped, ephemeral, and on-prem environments where the SaaS scanner cannot reach the target.",
     "/knoxctl/"),

    ("1: Technical Integration",
     "AccuKnox MCP Server",
     "Model Context Protocol server exposing AccuKnox findings and scan controls to LLM-based agents.",
     "The MCP server lets agents query DAST findings, trigger scans, and pull remediation guidance through a standard MCP interface. Enables LLM-driven triage and remediation copilots on top of AccuKnox data.",
     "/integrations/mcp-server/"),

    # 2: Various Scanning Types on DAST
    ("2: Various DAST Scanning Types",
     "Four-Tier DAST Scan Modes",
     "Baseline, Standard, Extended, and Comprehensive scan tiers controlling crawl depth, AJAX spider, and active-scan policy.",
     "Baseline is passive-only for every-commit gating. Standard runs the Dev CICD policy, Extended runs Dev Standard on SPAs, and Comprehensive runs Dev Full for pre-release audits. Tier selection drives both coverage and runtime predictability.",
     "/how-to/dast-scan-types/"),

    ("2: Various DAST Scanning Types",
     "Unauthenticated DAST Scan",
     "Black-box scanning of public surfaces with no credentials required.",
     "Point the scanner at a URL, pick a tier, and run. Used for marketing sites, public APIs, and first-pass coverage before authentication is wired up.",
     "/how-to/dast-scan-no-auth/"),

    ("2: Various DAST Scanning Types",
     "Authenticated DAST Scan",
     "Form login, session cookies, OAuth 2.0, JWT, and SSO-backed scans against private application surfaces.",
     "Auth is configured per target with multiple credential profiles, session validation rules, and a logged-in indicator to keep sessions alive across the scan.",
     "/how-to/dast-authenticated-scans/"),

    ("2: Various DAST Scanning Types",
     "MFA-Enabled DAST (TOTP)",
     "Scan targets behind TOTP-based multi-factor authentication without manual session handoff.",
     "The scanner accepts a TOTP secret and rotates codes during login. Removes the operational tax of running DAST against MFA-gated apps in staging or prod-like environments.",
     "/use-cases/mfa-dast/"),

    ("2: Various DAST Scanning Types",
     "XSS-Focused DAST Mitigation Flow",
     "Targeted scan profile and remediation playbook for cross-site scripting findings.",
     "Combines DAST detection with runtime KubeArmor policies to block reflected/stored XSS payloads at the kernel level until source-level fixes ship.",
     "/use-cases/dast-xss/"),

    # 3: Scanning from Collectors
    ("3: Scanning from Collectors",
     "Repo Collectors (GitHub, GitLab, Bitbucket)",
     "Lightweight collectors that pull source from SCM and run SCA, secrets, and IaC scans on every push.",
     "Repo Collectors run inside the customer environment and stream findings to the control plane. Source never leaves the customer network. Same pipeline feeds DAST scan triggers when paired with deployment events.",
     "/how-to/sca/"),

    ("3: Scanning from Collectors",
     "In-Cluster Image Scanner",
     "Helm-installed scanner that enumerates and scans every image running in a Kubernetes cluster.",
     "Discovers images at the kubelet level, scans for CVEs and malware, and reports without traffic leaving the cluster. Complements DAST by covering the binaries the application is built from.",
     "/how-to/in-cluster-image-scan-helm/"),

    ("3: Scanning from Collectors",
     "ECR Automated Scan Collector",
     "Auto-discovers new ECR repositories and tags, then scans every push without a manual onboarding step.",
     "Eliminates the per-repo registration overhead. New repos and new tags are picked up automatically and pushed to the same findings store as DAST.",
     "/how-to/ecr-automated-scan/"),

    ("3: Scanning from Collectors",
     "VM Container Image Scan via knoxctl",
     "Scan container images directly on a VM using knoxctl when no registry exists.",
     "Useful for edge nodes, disconnected sites, and bare-metal deployments where the image lives on disk rather than a registry.",
     "/knoxctl/image-scan/"),

    # 4: Runtime Security
    ("4: Runtime Security (Complements DAST)",
     "Kernel-Level Runtime Protection (KubeArmor)",
     "eBPF/LSM enforcement that blocks attacks pre-execution at the kernel level on Linux, K8s, and VMs.",
     "When DAST reports an exploitable path that cannot be fixed today, a KubeArmor policy blocks the matching syscall, file, or network action immediately, stopping the exploit while the code fix moves through the SDLC.",
     "/use-cases/cwpp/"),

    ("4: Runtime Security (Complements DAST)",
     "Runtime Security Architecture",
     "Reference architecture covering KubeArmor agents, relay, control plane, and policy distribution.",
     "Documents the data flow, system requirements, and policy lifecycle for AccuKnox runtime enforcement so SREs can size and harden the deployment.",
     "/getting-started/runtime-sec-arch/"),

    ("4: Runtime Security (Complements DAST)",
     "Workload Hardening Policies",
     "14+ pre-built hardening policies covering crypto miners, reverse shells, package manager abuse, and lateral movement.",
     "Drops a default-deny posture on the workload that turns DAST findings into a layered defense. Policies ship as code and version with the application.",
     "/use-cases/hardening/"),

    ("4: Runtime Security (Complements DAST)",
     "Network Micro-Segmentation",
     "Identity-aware egress/ingress control between services discovered automatically from runtime traffic.",
     "Reduces blast radius for any vulnerable endpoint DAST surfaces. Service-to-service paths are whitelisted; everything else is denied at the data plane.",
     "/use-cases/network-segmentation/"),

    ("4: Runtime Security (Complements DAST)",
     "KnoxGuard Admission Control",
     "Cluster admission controller that blocks workloads failing image, IaC, or policy checks before they run.",
     "DAST findings on a vulnerable image can feed a KnoxGuard rule that prevents the same image tag from ever reaching prod.",
     "/use-cases/admission-controller-knoxguard/"),

    ("4: Runtime Security (Complements DAST)",
     "Forensics / Audit Trail",
     "Per-process, per-syscall trace of what the application actually did during and after an attack.",
     "When DAST detects a high-severity finding, the runtime audit log shows whether the exploit succeeded, what files it touched, and what it tried to exfiltrate.",
     "/use-cases/forensics/"),

    # 5: Specification File Scanning
    ("5: Specification File Scanning",
     "OpenAPI / Swagger Upload",
     "Drive API security scans directly from an uploaded OpenAPI 2.x/3.x or Swagger spec.",
     "AccuKnox enumerates every endpoint defined in the spec, generates payloads per parameter, and tests authentication and authorization. Used for pre-prod APIs not yet exposed to live traffic.",
     "/integrations/api-overview/"),

    ("5: Specification File Scanning",
     "Postman Collection Import",
     "Run authenticated DAST against APIs defined in a Postman collection.",
     "Variables, auth headers, and request chains in the collection are honored. Removes the gap between what QA tests and what AccuKnox scans.",
     "/integrations/api-overview/"),

    ("5: Specification File Scanning",
     "cURL-Based Scan Onboarding",
     "Single cURL command captured from the browser dev tools is enough to onboard an endpoint.",
     "Lower-friction option for ad-hoc tests of a single endpoint without a full spec.",
     "/integrations/api-overview/"),

    ("5: Specification File Scanning",
     "Spec vs Runtime Drift Detection",
     "Compares the uploaded API spec against live traffic to surface undocumented and removed endpoints.",
     "Catches Shadow APIs (in traffic, not in spec), Zombie APIs (in spec, deprecated in traffic), and Orphan APIs (in code, not in any spec).",
     "/use-cases/api-security/"),

    # 6: API Security (extended)
    ("6: API Security (Extended)",
     "Runtime API Discovery via Gateway Traffic",
     "Continuous API inventory built from live request/response telemetry across gateways and service meshes.",
     "Inventory updates in real time without a pre-supplied spec. Each endpoint gets a method, params, auth posture, and data-class tag.",
     "/integrations/api-overview/"),

    ("6: API Security (Extended)",
     "AWS API Gateway Connector",
     "Native ingestion of API traffic and config from AWS API Gateway.",
     "Maps every Gateway resource and stage to the AccuKnox API inventory, with auth posture and rate-limit visibility.",
     "/integrations/api-aws/"),

    ("6: API Security (Extended)",
     "Kubernetes API Sec Proxy",
     "Sidecar/ingress-tap proxy that surfaces API traffic from any K8s workload without code changes.",
     "Deploys via Helm and routes mirrored traffic to the AccuKnox API security engine for inventory and runtime checks.",
     "/integrations/api-k8s/"),

    ("6: API Security (Extended)",
     "Istio Service Mesh Integration",
     "Pulls API telemetry directly from the Istio data plane.",
     "Reuses existing mesh sidecars so onboarding adds zero new proxies to the request path.",
     "/integrations/api-istio/"),

    ("6: API Security (Extended)",
     "Nginx Ingress Integration",
     "API observability via Nginx Ingress access logs and ingress controller hooks.",
     "Covers clusters that standardized on Nginx instead of a mesh.",
     "/integrations/api-nginx/"),

    ("6: API Security (Extended)",
     "Kong API Gateway Integration",
     "Plugin-based integration with Kong for inventory and runtime checks.",
     "Reads request metadata via a Kong plugin, no extra hop in the request path.",
     "/integrations/kong/"),

    ("6: API Security (Extended)",
     "F5 BIG-IP Integration",
     "Ingest traffic and config from F5 BIG-IP to extend API security into the data center perimeter.",
     "Covers hybrid topologies where F5 fronts north-south traffic before requests reach K8s or cloud.",
     "/integrations/f5/"),

    ("6: API Security (Extended)",
     "PII / PHI Classification on API Bodies",
     "Data-class tagging on request and response payloads at runtime.",
     "Flags endpoints leaking emails, SSNs, card numbers, PHI, and tokens so DAST findings on those endpoints get priority.",
     "/use-cases/api-security/"),

    ("6: API Security (Extended)",
     "Shadow / Zombie / Orphan API Detection",
     "Surface undocumented, deprecated-but-live, and code-only API endpoints.",
     "Cross-references uploaded specs, runtime traffic, and source-code routes to expose APIs the security team did not know existed.",
     "/integrations/api-overview/"),

    ("6: API Security (Extended)",
     "API Security FAQs",
     "Customer-facing FAQ covering API onboarding, supported protocols, and runtime modes.",
     "Operational reference for sales engineering and customer onboarding.",
     "/faqs/api-sec/"),

    # 7: Remediation
    ("7: Remediation",
     "AskADA GenAI Remediation Copilot",
     "Click any finding to get an AI-generated, contextual fix with code samples and verification steps.",
     "Works across DAST, SAST, SCA, IaC, and container findings. Supports batch remediation across many findings of the same class.",
     "/getting-started/3.3-release/"),

    ("7: Remediation",
     "Findings Lifecycle Management",
     "Workflow for triaging findings through New, In-Progress, Fixed, Accepted Risk, and False Positive states.",
     "Status changes drive auto-tickets, Slack alerts, and SIEM events without manual intervention.",
     "/how-to/findings-lifecycle/"),

    ("7: Remediation",
     "EPSS + CISA KEV Prioritization",
     "Risk score combining EPSS exploit probability, CISA KEV inclusion, CWE classification, and business impact weighting.",
     "DAST findings sort by real-world exploitability rather than CVSS alone, so the team fixes what actually gets exploited first.",
     "/use-cases/epss-scoring/"),

    ("7: Remediation",
     "Unified Vulnerability Management",
     "Single view across DAST, SAST, SCA, container, IaC, and cloud findings with deduplication and ownership tagging.",
     "Cuts triage time by eliminating the cross-tool reconciliation step that ASPM tools normally require.",
     "/use-cases/vulnerability/"),

    # 8: Automated Workflows / Rules Engine
    ("8: Automated Workflows",
     "Rules Engine (Condition-Based Actions)",
     "If/then rules that auto-create tickets, change finding status, post to Slack, or forward to SIEM based on any field.",
     "Common patterns: auto-ticket every Critical with EPSS > 0.7, suppress Info findings from non-prod, escalate KEV-listed CVEs to SOC.",
     "/use-cases/rules-engine-ticket-creation/"),

    ("8: Automated Workflows",
     "AI Natural-Language Rule Creation",
     "Describe a rule in plain English and the engine generates the corresponding condition tree.",
     "Lowers the bar for non-engineering teams to own automation. Generated rules are editable as structured conditions before activation.",
     "/use-cases/rules-engine-ticket-creation/"),

    ("8: Automated Workflows",
     "Smart Parent / Child Ticketing",
     "Aggregates many findings of the same root cause under a single parent ticket with child tickets per affected asset.",
     "Prevents ticket-storms during a release week. Closing the parent closes children automatically when the underlying CVE is resolved.",
     "/use-cases/rules-engine-ticket-creation/"),

    ("8: Automated Workflows",
     "Webhook Triggers",
     "Outbound webhooks fired on finding creation, status change, scan completion, and policy violation.",
     "Glue layer for any internal system that does not have a pre-built integration.",
     "/integrations/webhook-integration/"),

    # 9: Ticketing Workflows
    ("9: Ticketing Workflows",
     "Ticket Templates",
     "Reusable ticket bodies with merge fields for severity, asset, finding, and remediation.",
     "Lets each team apply its own ticket schema without code changes.",
     "/integrations/ticket-template/"),

    ("9: Ticketing Workflows",
     "Jira Cloud Integration",
     "Bi-directional sync with Jira Cloud for issue creation, status, and comments.",
     "Custom fields, project mapping, and per-severity assignee routing supported out of the box.",
     "/integrations/jira-cloud/"),

    ("9: Ticketing Workflows",
     "Jira Server (Self-Hosted) Integration",
     "Same workflow as Jira Cloud for on-prem Jira Data Center / Server.",
     "Required for regulated customers running Atlassian entirely on-prem.",
     "/integrations/jira-server-cspm/"),

    ("9: Ticketing Workflows",
     "ServiceNow Integration",
     "Create and sync incidents, problems, or change requests in ServiceNow ITSM.",
     "Field mapping covers assignment group, CI, urgency/impact, and free-form work notes.",
     "/integrations/servicenow/"),

    ("9: Ticketing Workflows",
     "Freshservice Integration",
     "Ticket creation and sync against Freshservice projects and queues.",
     "Targets mid-market customers standardized on the Freshworks suite.",
     "/integrations/freshservice-cspm/"),

    ("9: Ticketing Workflows",
     "ConnectWise Integration",
     "Native ticket sync for MSPs and MSSPs running ConnectWise PSA.",
     "Used by partners delivering AccuKnox as a managed service to downstream customers.",
     "/integrations/connectwise-cspm/"),

    ("9: Ticketing Workflows",
     "ManageEngine ServiceDesk Plus",
     "Ticket sync into ServiceDesk Plus on-prem or cloud.",
     "Covers enterprises standardized on the ManageEngine stack.",
     "/integrations/servicedesk-plus/"),

    # 10: ASPM Correlation & xBOM
    ("10: ASPM Correlation & xBOM",
     "Unified ASPM Across Scan Types",
     "Code-to-runtime correlation across DAST, SAST, SCA, IaC, container, secrets, and cloud findings.",
     "Same CVE found by DAST and SCA collapses into one finding with two evidence points. Risk score factors in runtime exposure.",
     "/use-cases/aspm/"),

    ("10: ASPM Correlation & xBOM",
     "xBOM (SBOM + CBOM + AIBOM)",
     "Generates Software, Cloud, and AI Bills of Materials in CycloneDX 1.6 and SPDX.",
     "BOM versions diff against each other, exposing newly introduced vulnerable components per release. Cosign signing supported for build attestation.",
     "/getting-started/xbom-setup/"),

    ("10: ASPM Correlation & xBOM",
     "ASPM Reports",
     "Pre-built and custom reports across all ASPM scan sources.",
     "Used for board reviews, audit evidence, and per-application security posture summaries.",
     "/use-cases/aspm-reports/"),

    # 11: Notifications, SIEM, Reporting
    ("11: Notifications, SIEM, Reporting",
     "Slack Notifications",
     "Route findings, scan summaries, and rule-engine alerts to Slack channels.",
     "Per-channel filters by severity, asset tag, or finding source.",
     "/integrations/slack/"),

    ("11: Notifications, SIEM, Reporting",
     "Email Notifications",
     "Scheduled and event-driven emails to security, DevOps, and compliance distribution lists.",
     "Configurable digest intervals to avoid alert fatigue.",
     "/integrations/email/"),

    ("11: Notifications, SIEM, Reporting",
     "Splunk Integration + AccuKnox Splunk App",
     "Forward DAST and runtime events to Splunk; AccuKnox Splunk App provides pre-built dashboards.",
     "Includes a feeder for KubeArmor telemetry so runtime context lands alongside DAST findings in the same Splunk indexes.",
     "/integrations/splunk/"),

    ("11: Notifications, SIEM, Reporting",
     "IBM QRadar Integration",
     "Push DAST and posture events to QRadar via webhook and syslog.",
     "Required by enterprises where QRadar is the SOC system of record.",
     "/integrations/ibm-qradar/"),

    ("11: Notifications, SIEM, Reporting",
     "Azure Sentinel Integration",
     "Send DAST, API, and runtime telemetry to Microsoft Sentinel.",
     "Native feeder plus syslog forwarding so events normalize into Sentinel analytics rules.",
     "/integrations/azure-sentinel/"),

    ("11: Notifications, SIEM, Reporting",
     "Rsyslog Forwarder",
     "RFC-5424 syslog forwarder for any SIEM not on the integration list.",
     "Catches the long tail (Graylog, Elastic, Chronicle) without custom code.",
     "/integrations/rsyslog/"),

    ("11: Notifications, SIEM, Reporting",
     "Custom Reports",
     "Build reports with selectable fields, filters, and date ranges across DAST and API findings.",
     "Reports export to PDF, XLSX, JSON, and can be scheduled to run and email automatically.",
     "/how-to/custom-reports/"),

    ("11: Notifications, SIEM, Reporting",
     "Summarized Custom Reports",
     "Executive-style summary reports designed for CISO and audit consumption.",
     "Aggregates DAST and API security posture into a one-page view with trendlines.",
     "/how-to/summarized-custom-reports/"),

    ("11: Notifications, SIEM, Reporting",
     "CNAPP Dashboard Widgets",
     "Drag-and-drop widgets for DAST and API security posture inside the CNAPP dashboard.",
     "Each persona (CISO, DevSecOps, Compliance, SOC) gets a default layout that can be customized per user.",
     "/use-cases/cnapp-security-overview/"),
]


def make_url(path: str) -> str:
    return f"{BASE}{path}"


def main():
    if not os.path.exists(SRC):
        raise SystemExit(f"Source not found: {SRC}")

    wb = load_workbook(SRC)

    sheet_name = "Appendix Additional DAST and API Security Features"
    # Excel sheet names are limited to 31 chars. Use a short form for the tab.
    short_name = "Appendix - Additional DAST_API"
    assert len(short_name) <= 31, len(short_name)

    # Replace if it already exists
    if short_name in wb.sheetnames:
        del wb[short_name]

    ws = wb.create_sheet(short_name)
    ws.sheet_properties.tabColor = "20124D"

    # Styles matching the source tab
    header_fill = PatternFill(fill_type="solid", fgColor="20124D")
    header_font = Font(name="Calibri", size=12, bold=True, color="FFFFFFFF")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    cat_fill = PatternFill(fill_type="solid", fgColor="D9D2E9")  # light purple
    cat_font = Font(name="Calibri", size=11, bold=True, color="FF20124D")
    cat_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    body_font = Font(name="Calibri", size=11)
    body_align_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)
    body_align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    link_font = Font(name="Calibri", size=11, color="FF0563C1", underline="single")

    thin = Side(border_style="thin", color="FFBFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Title row
    ws.cell(row=1, column=1).value = "Appendix - Additional AccuKnox DAST & API Security Features (Beyond RFP Scope)"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    tc = ws.cell(row=1, column=1)
    tc.fill = header_fill
    tc.font = Font(name="Calibri", size=14, bold=True, color="FFFFFFFF")
    tc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 32

    # Sub-title / context
    ws.cell(row=2, column=1).value = (
        "These rows highlight AccuKnox-native DAST and API security capabilities that were "
        "not represented in the original RFP requirement set. Every entry links to the public "
        "AccuKnox documentation at help.accuknox.com."
    )
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)
    sc = ws.cell(row=2, column=1)
    sc.font = Font(name="Calibri", size=10, italic=True, color="FF555555")
    sc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 36

    # Header row at row 3
    headers = ["Sr. No / Category", "Feature", "Description", "Type", "AccuKnox Capability", "Help Docs Reference"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=3, column=i, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = header_align
        c.border = border
    ws.row_dimensions[3].height = 30

    # Column widths matching the source tab where useful
    widths = {1: 27.86, 2: 30, 3: 45, 4: 18, 5: 55, 6: 50}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    # Write data rows starting row 4. Track category merges.
    current_row = 4
    cat_start = None
    last_cat = None
    sr_counter = 0

    def close_merge(end_row):
        nonlocal cat_start
        if cat_start is not None and end_row > cat_start:
            ws.merge_cells(start_row=cat_start, start_column=1, end_row=end_row, end_column=1)

    for category, feature, desc, capability, doc_path in ROWS:
        if category != last_cat:
            # Close previous merge
            close_merge(current_row - 1)
            # Start new merge on this row
            cat_start = current_row
            last_cat = category
            # Write category label in column A
            ca = ws.cell(row=current_row, column=1, value=category)
            ca.fill = cat_fill
            ca.font = cat_font
            ca.alignment = cat_align
            ca.border = border
        else:
            ca = ws.cell(row=current_row, column=1)
            ca.fill = cat_fill
            ca.border = border

        sr_counter += 1

        # B Feature
        c = ws.cell(row=current_row, column=2, value=feature)
        c.font = Font(name="Calibri", size=11, bold=True)
        c.alignment = body_align_wrap
        c.border = border

        # C Description
        c = ws.cell(row=current_row, column=3, value=desc)
        c.font = body_font
        c.alignment = body_align_wrap
        c.border = border

        # D Type
        c = ws.cell(row=current_row, column=4, value="Additional AccuKnox Capability")
        c.font = body_font
        c.alignment = body_align_center
        c.border = border

        # E AccuKnox Capability
        c = ws.cell(row=current_row, column=5, value=capability)
        c.font = body_font
        c.alignment = body_align_wrap
        c.border = border

        # F Help Docs Reference (hyperlink)
        url = make_url(doc_path)
        c = ws.cell(row=current_row, column=6, value=url)
        c.hyperlink = url
        c.font = link_font
        c.alignment = body_align_wrap
        c.border = border

        # Row height: enough to show ~3-5 lines of wrapped text
        ws.row_dimensions[current_row].height = 90

        current_row += 1

    # Close the final merge
    close_merge(current_row - 1)

    # Freeze the header row
    ws.freeze_panes = "B4"

    wb.save(DST)
    print(f"Saved appendix to: {DST}")
    print(f"Tab name: {short_name}")
    print(f"Rows: {len(ROWS)}")


if __name__ == "__main__":
    main()
