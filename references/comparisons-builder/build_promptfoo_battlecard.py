"""Build the AccuKnox vs Promptfoo AI security battlecard (.xlsx).

Run with:  py -3.11 build_promptfoo_battlecard.py

Every status value in ROWS is backed by a URL in the same row. Promptfoo claims
come from promptfoo.dev pages scraped 2026-08-15. AccuKnox claims come from
help.accuknox.com or accuknox.com.
"""

import os

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# AccuKnox brand palette (from the real proposal template)
NAVY = "11206D"
DEEP = "0000A0"
ACCENT = "4D4DD9"
DARK = "0D1B4B"
MID = "595959"
BORDER = "C4CCDE"
BAND = "F5F6FF"
SOLUTION = "EEF3FF"
WHITE = "FFFFFF"

YES_FILL = "DDF3E4"
YES_FONT = "1B6B3A"
NO_FILL = "FBE3E5"
NO_FONT = "A11220"
PART_FILL = "FFF3D6"
PART_FONT = "8A5A00"
BETA_FILL = "E3EEFF"
BETA_FONT = "1B4FA8"

HEAD = "Space Grotesk"
BODY = "Inter"

thin = Side(style="thin", color=BORDER)
BOX = Border(left=thin, right=thin, top=thin, bottom=thin)

# ---------------------------------------------------------------- source URLs
AK = {
    "shadow": "https://help.accuknox.com/use-cases/shadow-ai-discovery/",
    "aws_onb": "https://help.accuknox.com/how-to/aiml-aws-onboard/",
    "azure_onb": "https://help.accuknox.com/how-to/aiml-azure-onboard/",
    "gcp_onb": "https://help.accuknox.com/how-to/aiml-gcp-onboard/",
    "vm_linux": "https://help.accuknox.com/how-to/vm-security/agent-based/linux/",
    "vm_win": "https://help.accuknox.com/how-to/vm-security/agent-based/windows/",
    "k8s": "https://help.accuknox.com/how-to/k8s-security-onboarding/",
    "chrome": "https://help.accuknox.com/integrations/chrome-browser-integration/",
    "aispm": "https://help.accuknox.com/how-to/aiml-overview/",
    "platform_ai": "https://accuknox.com/platform/ai-security",
    "usecases": "https://help.accuknox.com/use-cases/aiml-usecases/",
    "ml_scan": "https://help.accuknox.com/how-to/ml-static-scan/",
    "llm_scan": "https://help.accuknox.com/how-to/llm-static-scan/",
    "modelarmor": "https://help.accuknox.com/use-cases/modelarmor/",
    "pickle": "https://help.accuknox.com/use-cases/modelarmor-pickle-code/",
    "redteam": "https://help.accuknox.com/use-cases/red-teaming/",
    "probes": "https://help.accuknox.com/use-cases/subprompts-categories/",
    "pf_over": "https://help.accuknox.com/use-cases/prompt-firewall-overview/",
    "pf_sdk": "https://help.accuknox.com/how-to/llm-defense-app-onboard/",
    "ai_integ": "https://help.accuknox.com/integrations/ai-overview/",
    "litellm": "https://help.accuknox.com/integrations/litellm/",
    "bifrost": "https://help.accuknox.com/integrations/bifrost-integration/",
    "apim": "https://help.accuknox.com/how-to/aws-apim/",
    "copilot": "https://help.accuknox.com/integrations/copilot-studio/",
    "agentcore": "https://help.accuknox.com/integrations/bedrock-agentcore/",
    "aidr": "https://help.accuknox.com/use-cases/aidr/",
    "agentic": "https://accuknox.com/solutions/agentic-ai-security",
    "agentz": "https://accuknox.com/platform/agentz",
    "agentz_gh": "https://github.com/accuknox/agentZ",
    "onprem": "https://help.accuknox.com/how-to/aiml-saas-vs-onprem/",
    "matrix": "https://help.accuknox.com/support-matrix/aiml-support-matrix/",
    "aspm": "https://help.accuknox.com/use-cases/aspm/",
    "compliance": "https://help.accuknox.com/use-cases/compliance/",
    "cicd": "https://help.accuknox.com/integrations/cicd-overview/",
    "mcp_srv": "https://help.accuknox.com/integrations/mcp-server/",
    "rules": "https://help.accuknox.com/use-cases/rules-engine-ticket-creation/",
    "faq": "https://help.accuknox.com/faqs/ai-security/",
    "vllm": "https://help.accuknox.com/how-to/aiml-vllm-collector/",
    "triton": "https://help.accuknox.com/how-to/aiml-triton-collector/",
    "servicenow": "https://help.accuknox.com/integrations/servicenow/",
    "integ_matrix": "https://help.accuknox.com/integrations/support-matrix/",
    "cnapp": "https://help.accuknox.com/use-cases/cnapp-security-overview/",
}

PF = {
    "home": "https://www.promptfoo.dev/",
    "redteam": "https://www.promptfoo.dev/red-teaming/",
    "plugins": "https://www.promptfoo.dev/docs/red-team/plugins/",
    "agents": "https://www.promptfoo.dev/docs/red-team/agents/",
    "multiturn": "https://www.promptfoo.dev/docs/red-team/strategies/multi-turn/",
    "multimodal": "https://www.promptfoo.dev/docs/guides/multimodal-red-team/",
    "model_sec": "https://www.promptfoo.dev/model-security/",
    "guardrails": "https://www.promptfoo.dev/guardrails/",
    "gr_assert": "https://www.promptfoo.dev/docs/configuration/expected-outputs/guardrails/",
    "mcp": "https://www.promptfoo.dev/mcp/",
    "mcp_test": "https://www.promptfoo.dev/docs/red-team/mcp-security-testing/",
    "code": "https://www.promptfoo.dev/code-scanning/",
    "vscode": "https://www.promptfoo.dev/docs/code-scanning/vscode-extension/",
    "cicd": "https://www.promptfoo.dev/docs/integrations/ci-cd/",
    "selfhost": "https://www.promptfoo.dev/docs/usage/self-hosting/",
    "pricing": "https://www.promptfoo.dev/pricing/",
    "owasp": "https://www.promptfoo.dev/docs/red-team/owasp-llm-top-10/",
    "nist": "https://www.promptfoo.dev/docs/red-team/nist-ai-rmf/",
    "euai": "https://www.promptfoo.dev/docs/red-team/eu-ai-act/",
    "audit": "https://www.promptfoo.dev/docs/enterprise/audit-logging/",
    "teams": "https://www.promptfoo.dev/docs/enterprise/teams/",
    "tracing": "https://www.promptfoo.dev/docs/tracing/",
    "scanner": "https://www.promptfoo.dev/llm-vulnerability-scanner/",
    "mcp_srv": "https://www.promptfoo.dev/docs/integrations/mcp-server/",
}

NOPAGE = ("Not offered. Their six products are Red Teaming, Evaluations, Code "
          "Scanning, Model Security, Guardrails, MCP: promptfoo.dev")

# ------------------------------------------------------------------ the table
# (module, sub-feature, accuknox, promptfoo, ak_ref, pf_ref)
ROWS = [
    ("1. Shadow AI Discovery",
     "Agentless AI asset discovery across AWS, Azure, GCP",
     "Yes", "No", AK["shadow"] + " | " + AK["aws_onb"], NOPAGE),
    ("1. Shadow AI Discovery",
     "On-prem VM discovery: package-level scan of the host filesystem",
     "Yes", "No", AK["shadow"] + " | " + AK["vm_linux"], NOPAGE),
    ("1. Shadow AI Discovery",
     "In-cluster Kubernetes discovery across nodes, namespaces, images",
     "Yes", "No", AK["shadow"] + " | " + AK["k8s"], NOPAGE),
    ("1. Shadow AI Discovery",
     "Classify assets into 7 types: Agent, Automation, Gateway, Inference "
     "Engine, AI-ML, SDK, MCP",
     "Yes", "No", AK["shadow"], NOPAGE),
    ("1. Shadow AI Discovery",
     "Find unapproved agentic runtimes installed on workstations and build VMs",
     "Yes", "No", AK["shadow"], NOPAGE),
    ("1. Shadow AI Discovery",
     "Browser plugin that records GenAI service usage (ChatGPT, Claude, Gemini, "
     "Copilot)",
     "Yes", "No", AK["chrome"] + " | " + AK["shadow"], NOPAGE),
    ("1. Shadow AI Discovery",
     "Cloud AI pipeline graph: connectivity, exposure, service relationships",
     "Yes", "No", AK["shadow"], NOPAGE),
    ("1. Shadow AI Discovery",
     "Vulnerability and malware findings attached to each discovered AI asset",
     "Yes", "No", AK["shadow"], NOPAGE),

    ("2. AI Posture Management (AI-SPM)",
     "Live inventory of models, agents, datasets, and pipelines in one console",
     "Yes", "No", AK["aispm"] + " | " + AK["platform_ai"], NOPAGE),
    ("2. AI Posture Management (AI-SPM)",
     "Misconfiguration posture for cloud AI services (Bedrock, SageMaker, "
     "Azure OpenAI, Vertex)",
     "Yes", "No", AK["aispm"] + " | " + AK["platform_ai"], NOPAGE),
    ("2. AI Posture Management (AI-SPM)",
     "Auto-remediation of publicly exposed AI endpoints",
     "Yes", "No", AK["aidr"] + " | " + AK["platform_ai"], NOPAGE),
    ("2. AI Posture Management (AI-SPM)",
     "Security graph that maps agents to models, tools, datasets, cloud accounts",
     "Yes", "No", AK["shadow"] + " | " + AK["aispm"], NOPAGE),
    ("2. AI Posture Management (AI-SPM)",
     "AI Bill of Materials (AIBOM) generation",
     "Yes", "No", AK["platform_ai"], NOPAGE),
    ("2. AI Posture Management (AI-SPM)",
     "Onboarding for self-hosted inference engines (vLLM, Triton, Bedrock)",
     "Yes", "No", AK["vllm"] + " | " + AK["triton"], NOPAGE),

    ("3. AI Model & Dataset Security",
     "Static scan of model files for malicious code and deserialization attacks",
     "Yes", "Yes", AK["ml_scan"], PF["model_sec"]),
    ("3. AI Model & Dataset Security",
     "Model formats covered",
     "Yes", "Yes",
     "5 formats — Pickle, HDF5/H5, TensorFlow SavedModel, Checkpoints, ONNX: "
     + AK["ml_scan"],
     "PyTorch, TensorFlow, Keras, Pickle, JSON/YAML: " + PF["model_sec"]),
    ("3. AI Model & Dataset Security",
     "Block a poisoned model from public repos before it deploys",
     "Yes", "Partial", AK["pickle"] + " | " + AK["platform_ai"],
     "Flags risky files, no deploy-time gate documented: " + PF["model_sec"]),
    ("3. AI Model & Dataset Security",
     "PII and PHI scanning of training datasets and vector embeddings",
     "Yes", "No", AK["faq"] + " (question 24)", NOPAGE),
    ("3. AI Model & Dataset Security",
     "Dataset lineage and data poisoning detection",
     "Yes", "No", AK["faq"] + " (question 24)", NOPAGE),
    ("3. AI Model & Dataset Security",
     "Runtime protection of model execution (eBPF / LSM)",
     "Yes", "No", AK["modelarmor"], NOPAGE),
    ("3. AI Model & Dataset Security",
     "Findings mapped to OWASP LLM Top 10 and MITRE ATLAS",
     "Yes", "Yes", AK["redteam"] + " | " + AK["platform_ai"], PF["model_sec"]),
    ("3. AI Model & Dataset Security",
     "Foundation model benchmarking, security score per model",
     "No", "Yes", "Not documented", PF["model_sec"]),

    ("4. AI Red Teaming & Pen Testing",
     "Automated adversarial probe library",
     "Yes", "Yes",
     "150+ probes across 4 categories: " + AK["redteam"] + " | " + AK["probes"],
     "157 plugins across 6 categories: " + PF["plugins"]),
    ("4. AI Red Teaming & Pen Testing",
     "Scans re-run on a schedule and after every model change",
     "Yes", "Yes", AK["redteam"], PF["redteam"]),
    ("4. AI Red Teaming & Pen Testing",
     "Attacks generated dynamically from your application context",
     "Partial", "Yes",
     "Domain-specific probe packs, custom prompt files: " + AK["llm_scan"],
     "Core capability, attacks built per target: " + PF["redteam"]),
    ("4. AI Red Teaming & Pen Testing",
     "Multi-turn conversational jailbreak strategies",
     "Partial", "Yes",
     "Multi-turn tracking sits in the Prompt Firewall, not the scanner: "
     + AK["pf_over"], PF["multiturn"]),
    ("4. AI Red Teaming & Pen Testing",
     "Public research datasets (HarmBench, BeaverTails, CyberSecEval, Aegis, "
     "Pliny)",
     "No", "Yes", "Not documented", PF["plugins"]),
    ("4. AI Red Teaming & Pen Testing",
     "Regulated-industry probe packs (finance, insurance, telecom, pharmacy, "
     "real estate)",
     "Partial", "Yes",
     "Custom domain probes, no shipped industry packs: " + AK["redteam"],
     PF["plugins"]),
    ("4. AI Red Teaming & Pen Testing",
     "Multimodal red teaming (image, audio, file inputs)",
     "No", "Yes", "Not documented", PF["multimodal"]),
    ("4. AI Red Teaming & Pen Testing",
     "Coding-agent red teaming (repo prompt injection, delayed CI exfiltration)",
     "No", "Yes", "Not documented", PF["plugins"]),
    ("4. AI Red Teaming & Pen Testing",
     "Agent and MCP adversarial testing (tool poisoning, rug pull, memory "
     "poisoning)",
     "Partial", "Yes",
     "Covered by runtime sandboxing rather than a test suite: " + AK["modelarmor"],
     PF["mcp_test"] + " | " + PF["agents"]),
    ("4. AI Red Teaming & Pen Testing",
     "Findings export and remediation workflow",
     "Yes", "Yes", AK["redteam"], PF["redteam"]),

    ("5. AI Guardrails (Prompt Firewall)",
     "Inline proxy that inspects every prompt and every response in production",
     "Yes", "Partial",
     "14 policy classes, block / sanitize / monitor: " + AK["pf_over"],
     "Enterprise adaptive guardrails, marketing page only; the OSS `guardrails` "
     "assert only reads Bedrock or Azure filter verdicts: " + PF["guardrails"]
     + " | " + PF["gr_assert"]),
    ("5. AI Guardrails (Prompt Firewall)",
     "PII and PHI masking in prompts and responses",
     "Yes", "Partial", AK["pf_over"], PF["guardrails"] + " | " + PF["gr_assert"]),
    ("5. AI Guardrails (Prompt Firewall)",
     "Secrets detection inside prompts",
     "Yes", "No", AK["pf_over"], NOPAGE),
    ("5. AI Guardrails (Prompt Firewall)",
     "Ban topics, ban competitors, ban code, language and regex policies",
     "Yes", "No", AK["pf_over"], NOPAGE),
    ("5. AI Guardrails (Prompt Firewall)",
     "Stateful multi-turn tracking across a live conversation",
     "Yes", "No", AK["pf_over"], NOPAGE),
    ("5. AI Guardrails (Prompt Firewall)",
     "AI gateway integrations (Azure APIM, AWS API Gateway, Apigee, LiteLLM, "
     "Bifrost)",
     "Yes", "No", AK["ai_integ"] + " | " + AK["litellm"] + " | " + AK["bifrost"],
     NOPAGE),
    ("5. AI Guardrails (Prompt Firewall)",
     "Enterprise app integrations (Copilot Studio, Power Apps, Bedrock "
     "AgentCore)",
     "Yes", "No", AK["copilot"] + " | " + AK["agentcore"], NOPAGE),
    ("5. AI Guardrails (Prompt Firewall)",
     "Browser-side guardrails (Chrome, Edge, Firefox)",
     "Yes", "No", AK["chrome"], NOPAGE),
    ("5. AI Guardrails (Prompt Firewall)",
     "Python SDK to protect an app directly",
     "Yes", "No",
     AK["pf_sdk"], "SDK exists for testing, not runtime enforcement: " + PF["home"]),
    ("5. AI Guardrails (Prompt Firewall)",
     "Test and validate a third-party guardrail",
     "Partial", "Yes", "Red teaming can probe an app behind any guardrail: "
     + AK["redteam"], PF["guardrails"]),

    ("6. Agentic AI & MCP Security",
     "Runtime sandbox for agents: process, file, network, credential isolation",
     "Yes", "No", AK["modelarmor"] + " | " + AK["agentic"], NOPAGE),
    ("6. Agentic AI & MCP Security",
     "MCP tool execution sandboxed with least-permissive access",
     "Yes", "No", AK["modelarmor"] + " | " + AK["agentic"], NOPAGE),
    ("6. Agentic AI & MCP Security",
     "MCP proxy that allowlists approved MCP servers org-wide",
     "No", "Yes", "Not documented", PF["mcp"]),
    ("6. Agentic AI & MCP Security",
     "MCP request logging with alerts on sensitive data exposure",
     "Partial", "Yes",
     "Covered by AI-DR and Prompt Firewall audit logs: " + AK["aidr"], PF["mcp"]),
    ("6. Agentic AI & MCP Security",
     "Per-agent cryptographic identity (SPIFFE)",
     "Roadmap", "No",
     "Described on accuknox.com; help docs list AI Identity Security as coming "
     "soon: " + AK["platform_ai"] + " | " + AK["usecases"], NOPAGE),
    ("6. Agentic AI & MCP Security",
     "Fine-grained per-agent authorization on tool calls",
     "Yes", "Partial", AK["platform_ai"] + " | " + AK["agentz"],
     "Access control on MCP servers, not on individual tool calls: " + PF["mcp"]),
    ("6. Agentic AI & MCP Security",
     "Memory poisoning — runtime prevention",
     "Yes", "No", AK["agentic"], NOPAGE),
    ("6. Agentic AI & MCP Security",
     "Memory poisoning — adversarial testing",
     "No", "Yes", "Not documented", PF["agents"]),
    ("6. Agentic AI & MCP Security",
     "Agent execution tracing (OpenTelemetry spans, tool calls)",
     "Partial", "Yes",
     "Trace and replay sit in AgentZ, see module 7: " + AK["agentz"],
     PF["tracing"]),

    ("7. Agentic Harness Platform (AgentZ)",
     "Build, run, schedule, and govern production agents on one control plane",
     "Yes", "No", AK["agentz"] + " | " + AK["agentz_gh"], NOPAGE),
    ("7. Agentic Harness Platform (AgentZ)",
     "Every agent runs in a default-deny Zero Trust sandbox from the first run",
     "Yes", "No", AK["agentz"], NOPAGE),
    ("7. Agentic Harness Platform (AgentZ)",
     "Network egress allowed or blocked at the kernel, by domain, port, protocol",
     "Yes", "No", AK["agentz"], NOPAGE),
    ("7. Agentic Harness Platform (AgentZ)",
     "Agents never hold secrets — scoped credentials injected at run time",
     "Yes", "No", AK["agentz"], NOPAGE),
    ("7. Agentic Harness Platform (AgentZ)",
     "RBAC down to the individual tool call, with roles and teams",
     "Yes", "No", AK["agentz"], NOPAGE),
    ("7. Agentic Harness Platform (AgentZ)",
     "Full replayable audit trace with a deterministic replay ID",
     "Yes", "No", AK["agentz"], NOPAGE),
    ("7. Agentic Harness Platform (AgentZ)",
     "Model independence — BYOK, BYOS, BYOM including private models",
     "Yes", "No", AK["agentz"], NOPAGE),
    ("7. Agentic Harness Platform (AgentZ)",
     "Wrap an existing LangGraph or CrewAI agent under the same policy edge",
     "Yes", "No", AK["agentz"], NOPAGE),
    ("7. Agentic Harness Platform (AgentZ)",
     "Air-gapped deployment with no outbound calls",
     "Yes", "No", AK["agentz"], NOPAGE),

    ("8. AI Detection & Response (AI-DR)",
     "Continuous ingest of AWS CloudTrail, Azure, and GCP control-plane logs",
     "Yes", "No", AK["aidr"], NOPAGE),
    ("8. AI Detection & Response (AI-DR)",
     "Detect unauthorized AI deployments and publicly exposed model endpoints",
     "Yes", "No", AK["aidr"], NOPAGE),
    ("8. AI Detection & Response (AI-DR)",
     "Automated response — make an exposed AI asset private again",
     "Yes", "No", AK["aidr"], NOPAGE),
    ("8. AI Detection & Response (AI-DR)",
     "Attack path reconstruction across prompt, model, API, and infrastructure",
     "Yes", "No", AK["aidr"] + " | " + AK["platform_ai"], NOPAGE),
    ("8. AI Detection & Response (AI-DR)",
     "Ticket and alert routing (Jira, ServiceNow, Slack, PagerDuty)",
     "Yes", "Partial",
     AK["servicenow"] + " | " + AK["integ_matrix"] + " | " + AK["rules"],
     "Webhooks on Enterprise, no named ITSM connector: " + PF["pricing"]),

    ("9. Compliance & Governance",
     "OWASP Top 10 for LLMs mapping",
     "Yes", "Yes", AK["redteam"], PF["owasp"]),
    ("9. Compliance & Governance",
     "MITRE ATLAS mapping",
     "Yes", "Yes", AK["redteam"] + " | " + AK["onprem"], PF["model_sec"]),
    ("9. Compliance & Governance",
     "NIST AI RMF mapping",
     "Yes", "Yes", AK["onprem"], PF["nist"]),
    ("9. Compliance & Governance",
     "EU AI Act mapping",
     "Yes", "Yes", AK["platform_ai"], PF["euai"]),
    ("9. Compliance & Governance",
     "AVID (AI Vulnerability Database) mapping",
     "Yes", "No", AK["onprem"], NOPAGE),
    ("9. Compliance & Governance",
     "Compliance beyond AI — CIS, PCI-DSS, SOC 2, HIPAA, ISO 27001, NIST",
     "Yes", "No", AK["compliance"], NOPAGE),
    ("9. Compliance & Governance",
     "Audit logging of platform activity",
     "Yes", "Yes", AK["aidr"], PF["audit"]),
    ("9. Compliance & Governance",
     "SSO, teams, and granular permission profiles",
     "Yes", "Yes", AK["mcp_srv"], PF["teams"]),

    ("10. Developer Workflow",
     "CI/CD pipeline integration (Jenkins, GitHub Actions, GitLab, Azure DevOps)",
     "Yes", "Yes", AK["cicd"], PF["cicd"]),
    ("10. Developer Workflow",
     "LLM-specific code scanning — data flow from user input into prompts",
     "No", "Yes", "General SAST, SCA, IaC and secrets only: " + AK["aspm"],
     PF["code"]),
    ("10. Developer Workflow",
     "IDE extension with inline findings (VS Code)",
     "No", "Yes", "Not documented", PF["vscode"]),
    ("10. Developer Workflow",
     "Security findings posted as pull request comments",
     "Partial", "Yes", "ASPM gates the pipeline, no PR review comments: "
     + AK["aspm"], PF["code"]),
    ("10. Developer Workflow",
     "Free open-source tier a developer can run locally today",
     "Partial", "Yes",
     "KubeArmor and the ASPM scanner CLI are OSS, the AI modules are not: "
     + AK["aspm"], PF["pricing"]),
    ("10. Developer Workflow",
     "Ships an MCP server so engineers can drive it from their AI tools",
     "Yes", "Yes",
     "Query assets, findings, and compliance: " + AK["mcp_srv"],
     "Exposes eval tools to agents: " + PF["mcp_srv"]),

    ("11. Deployment & Platform",
     "SaaS delivery",
     "Yes", "Yes", AK["onprem"], PF["pricing"]),
    ("11. Deployment & Platform",
     "Self-hosted on customer infrastructure",
     "Yes", "Yes",
     "Helm on self-hosted K8s or a 3-node VM cluster, full feature parity: "
     + AK["onprem"],
     "Docker, Docker Compose, or an experimental Helm chart: " + PF["selfhost"]),
    ("11. Deployment & Platform",
     "Air-gapped deployment",
     "Yes", "Partial", AK["onprem"],
     "On-prem tier is offered, air-gap is not documented: " + PF["pricing"]),
    ("11. Deployment & Platform",
     "Production-grade horizontal scaling of the self-hosted server",
     "Yes", "No", AK["onprem"],
     "SQLite backed, multiple replicas break job lookup: " + PF["selfhost"]),
    ("11. Deployment & Platform",
     "Scope beyond AI — CNAPP, CSPM, CWPP, KSPM, ASPM in one platform",
     "Yes", "No", AK["cnapp"] + " | " + AK["compliance"], NOPAGE),
    ("11. Deployment & Platform",
     "Independent vendor",
     "Yes", "No", "AccuKnox is independent: " + AK["platform_ai"],
     "Promptfoo states it is part of OpenAI, on every page banner: " + PF["home"]),
]

# --------------------------------------------------------------- module notes
MODULE_NOTES = {
    "7. Agentic Harness Platform (AgentZ)":
        "Side note — AgentZ is a separate AccuKnox product, not a module inside "
        "AI-SPM. It is the platform you build and run agents on, where the other "
        "modules secure agents you already run. Promptfoo has no equivalent: it "
        "tests agents, it does not host them.",
}

STATUS_STYLE = {
    "Yes": (YES_FILL, YES_FONT),
    "No": (NO_FILL, NO_FONT),
    "Partial": (PART_FILL, PART_FONT),
    "Beta": (BETA_FILL, BETA_FONT),
    "Roadmap": (BETA_FILL, BETA_FONT),
}


def style_status(cell, value):
    fill, font = STATUS_STYLE.get(value, (WHITE, DARK))
    cell.fill = PatternFill("solid", fgColor=fill)
    cell.font = Font(name=BODY, size=10, bold=True, color=font)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = BOX


def build():
    wb = Workbook()

    # ============================================================ SHEET 1
    ws = wb.active
    ws.title = "Stack Ranking"
    ws.sheet_view.showGridLines = False

    widths = [30, 62, 13, 13, 60, 60]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    r = 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    c = ws.cell(r, 1, "AI SECURITY STACK RANKING — AccuKnox vs Promptfoo")
    c.font = Font(name=HEAD, size=18, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[r].height = 38
    r += 1

    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    c = ws.cell(r, 1,
                "Every Yes, No, and Partial in this sheet has a URL in the same "
                "row. Promptfoo claims come from promptfoo.dev, read on "
                "15 August 2026. AccuKnox claims come from help.accuknox.com or "
                "accuknox.com. Where Promptfoo is stronger, the sheet says so.")
    c.font = Font(name=BODY, size=9, color=MID)
    c.fill = PatternFill("solid", fgColor=BAND)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1,
                            wrap_text=True)
    ws.row_dimensions[r].height = 30
    r += 1

    # scorecard
    ak_yes = sum(1 for x in ROWS if x[2] == "Yes")
    pf_yes = sum(1 for x in ROWS if x[3] == "Yes")
    both = sum(1 for x in ROWS if x[2] == "Yes" and x[3] == "Yes")
    ak_only = sum(1 for x in ROWS if x[2] == "Yes" and x[3] in ("No", "Partial"))
    pf_only = sum(1 for x in ROWS if x[3] == "Yes" and x[2] in ("No", "Partial"))

    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    c = ws.cell(r, 1,
                f"SCORECARD ({len(ROWS)} sub-features)     "
                f"AccuKnox full support: {ak_yes}     "
                f"Promptfoo full support: {pf_yes}     "
                f"Both: {both}     "
                f"AccuKnox only: {ak_only}     "
                f"Promptfoo only: {pf_only}")
    c.font = Font(name=HEAD, size=11, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=ACCENT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[r].height = 26
    r += 1

    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    c = ws.cell(r, 1,
                "LEGEND    Yes = shipped and documented    "
                "Partial = related capability, narrower than the row asks for    "
                "No = no public documentation    "
                "Beta = shipping, not yet GA    "
                "Roadmap = announced, not shipped")
    c.font = Font(name=BODY, size=9, italic=True, color=MID)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[r].height = 20
    r += 2

    headers = ["Module", "Sub-feature", "AccuKnox", "Promptfoo",
               "AccuKnox reference", "Promptfoo reference"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(r, i, h)
        c.font = Font(name=HEAD, size=10, bold=True, color=WHITE)
        c.fill = PatternFill("solid", fgColor=DEEP)
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)
        c.border = BOX
    ws.row_dimensions[r].height = 30
    ws.freeze_panes = ws.cell(r + 1, 1)
    r += 1

    last_module = None
    first_of_group = True
    for module, feature, ak, pf, ak_ref, pf_ref in ROWS:
        if module != last_module:
            # module banner
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
            c = ws.cell(r, 1, module.upper())
            c.font = Font(name=HEAD, size=10, bold=True, color=WHITE)
            c.fill = PatternFill("solid", fgColor=NAVY)
            c.alignment = Alignment(horizontal="left", vertical="center",
                                    indent=1)
            ws.row_dimensions[r].height = 22
            r += 1
            note = MODULE_NOTES.get(module)
            if note:
                ws.merge_cells(start_row=r, start_column=1, end_row=r,
                               end_column=6)
                c = ws.cell(r, 1, note)
                c.font = Font(name=BODY, size=9, italic=True, color=ACCENT)
                c.fill = PatternFill("solid", fgColor=SOLUTION)
                c.alignment = Alignment(horizontal="left", vertical="center",
                                        indent=1, wrap_text=True)
                ws.row_dimensions[r].height = 30
                r += 1
            last_module = module
            first_of_group = True

        band = BAND if (r % 2 == 0) else WHITE

        c = ws.cell(r, 1, module if first_of_group else "")
        first_of_group = False
        c.font = Font(name=BODY, size=9, color=MID)
        c.alignment = Alignment(vertical="center", wrap_text=True, indent=1)
        c.fill = PatternFill("solid", fgColor=band)
        c.border = BOX

        c = ws.cell(r, 2, feature)
        c.font = Font(name=BODY, size=10, color=DARK)
        c.alignment = Alignment(vertical="center", wrap_text=True, indent=1)
        c.fill = PatternFill("solid", fgColor=band)
        c.border = BOX

        style_status(ws.cell(r, 3, ak), ak)
        style_status(ws.cell(r, 4, pf), pf)

        for col, ref in ((5, ak_ref), (6, pf_ref)):
            c = ws.cell(r, col, ref)
            c.font = Font(name=BODY, size=8, color=MID)
            c.alignment = Alignment(vertical="center", wrap_text=True, indent=1)
            c.fill = PatternFill("solid", fgColor=band)
            c.border = BOX

        ws.row_dimensions[r].height = 34
        r += 1

    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    c = ws.cell(r, 1,
                "READ THIS BEFORE YOU USE THE SHEET — Promptfoo is an AI "
                "red-teaming and evaluation tool, not an AI security platform. "
                "It is excellent at what it does: 157 attack plugins, dynamic "
                "attack generation, LLM-aware code scanning, and an MCP proxy. "
                "It has no asset discovery, no cloud posture, no runtime "
                "enforcement, and no detection and response. Sell the gap, not "
                "a like-for-like win.")
    c.font = Font(name=BODY, size=10, bold=True, color=DARK)
    c.fill = PatternFill("solid", fgColor=SOLUTION)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1,
                            wrap_text=True)
    ws.row_dimensions[r].height = 46

    # ============================================================ SHEET 2
    ws2 = wb.create_sheet("Positioning")
    ws2.sheet_view.showGridLines = False
    for col, w in zip("ABC", (34, 78, 78)):
        ws2.column_dimensions[col].width = w

    r = 1
    ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    c = ws2.cell(r, 1, "POSITIONING — WHERE EACH SIDE ACTUALLY WINS")
    c.font = Font(name=HEAD, size=16, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(vertical="center", indent=1)
    ws2.row_dimensions[r].height = 34
    r += 2

    blocks = [
        ("Where AccuKnox wins",
         "Everything that happens outside a test run. You cannot test an asset "
         "you have not found, and Promptfoo has no discovery at all. Shadow AI "
         "Discovery, AI-SPM inventory, cloud misconfiguration posture, AI-DR on "
         "CloudTrail and Azure logs, eBPF runtime sandboxing of agents and MCP "
         "tools, the inline Prompt Firewall with 14 policy classes, SPIFFE agent "
         "identity, dataset PII scanning, and AgentZ as a governed agent "
         "harness. Add air-gapped deployment with full feature parity and the "
         "rest of the CNAPP platform underneath.",
         "Ask: how do you secure the AI you do not know you are running?"),
        ("Where Promptfoo wins",
         "Depth of offensive testing. 157 attack plugins across 6 categories, "
         "including public research datasets and regulated-industry packs we do "
         "not ship. Dynamic attacks generated per target rather than a fixed "
         "probe list. Multimodal and coding-agent red teaming. LLM-aware code "
         "scanning that traces untrusted input into prompts, delivered in the "
         "IDE and in pull request comments. A free open-source tier a developer "
         "installs in one command, and an MCP proxy that allowlists servers.",
         "Do not argue these. Concede them, then move the conversation to "
         "runtime and coverage."),
        ("The honest overlap",
         "Both scan model files, both map findings to OWASP LLM Top 10, NIST AI "
         "RMF, and the EU AI Act, both run in CI/CD, both self-host, and both "
         "offer SSO and audit logging. Do not claim a win on any of these rows.",
         "Parity rows build the credibility that carries the gap rows."),
        ("Vendor and buying context",
         "Promptfoo states on every page that it is part of OpenAI. For a BFSI, "
         "government, or PSU buyer running a multi-model estate, that is a "
         "procurement question worth raising, factually and without spin. Their "
         "self-hosted server is SQLite backed and documented as unable to run "
         "multiple replicas, and the Helm chart is marked experimental.",
         "Source: promptfoo.dev banner, and the self-hosting doc."),
        ("Where they might beat us in a bake-off",
         "A developer-led evaluation that starts with 'run a red team scan "
         "tonight' favours Promptfoo. They install free in one command and "
         "produce a report in about 14 minutes by their own claim. Get the "
         "evaluation scoped to the full AI estate, not to one chatbot.",
         "Counter: run our Shadow AI scan on their cloud account in parallel and "
         "show them assets they did not list."),
    ]
    for title, body, tag in blocks:
        c = ws2.cell(r, 1, title)
        c.font = Font(name=HEAD, size=11, bold=True, color=WHITE)
        c.fill = PatternFill("solid", fgColor=ACCENT)
        c.alignment = Alignment(vertical="center", indent=1, wrap_text=True)
        c.border = BOX
        c2 = ws2.cell(r, 2, body)
        c2.font = Font(name=BODY, size=10, color=DARK)
        c2.alignment = Alignment(vertical="top", wrap_text=True, indent=1)
        c2.border = BOX
        c3 = ws2.cell(r, 3, tag)
        c3.font = Font(name=BODY, size=10, italic=True, color=MID)
        c3.alignment = Alignment(vertical="top", wrap_text=True, indent=1)
        c3.border = BOX
        ws2.row_dimensions[r].height = 96
        r += 1

    # ============================================================ SHEET 3
    ws3 = wb.create_sheet("Proof Screens")
    ws3.sheet_view.showGridLines = False
    ws3.column_dimensions["A"].width = 4
    ws3.column_dimensions["B"].width = 120

    r = 1
    ws3.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    c = ws3.cell(r, 1, "PROOF SCREENS — AccuKnox product evidence")
    c.font = Font(name=HEAD, size=16, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(vertical="center", indent=1)
    ws3.row_dimensions[r].height = 34
    r += 2

    docs = r"D:\Atharva\AccuKnox\HelpDocs\docs"
    shots = [
        (os.path.join(docs, "use-cases", "images", "shadow-ai",
                      "unmanaged-asset-categories.png"),
         "Shadow AI Discovery — on-prem assets rolled up into the seven AI/ML "
         "categories, each with an asset count and findings by severity. "
         "Promptfoo has no screen like this.",
         AK["shadow"], 900),
        (os.path.join(docs, "use-cases", "images", "shadow-ai",
                      "managed-assets-models.png"),
         "AI-SPM — cloud-discovered models on the Managed tab, with cloud type, "
         "region, and last seen date.",
         AK["shadow"], 900),
        (os.path.join(docs, "use-cases", "image-1.png"),
         "Prompt Firewall — the inline proxy, the policy governance engine, and "
         "the audit log sitting between client apps and the model provider.",
         AK["pf_over"], 900),
        (os.path.join(docs, "use-cases", "images", "aidr", "1.png"),
         "AI-DR — cloud control-plane events flow into policy evaluation, "
         "alerting, ticketing, and automated remediation.",
         AK["aidr"], 820),
        (os.path.join(docs, "use-cases", "images", "modelarmor", "1.png"),
         "ModelArmor — KubeArmor sandboxes agent and model execution at "
         "runtime, enforcing process, file, network, and credential isolation.",
         AK["modelarmor"], 820),
    ]

    for path, caption, url, target_w in shots:
        if not os.path.exists(path):
            print("MISSING IMAGE:", path)
            continue
        c = ws3.cell(r, 2, caption)
        c.font = Font(name=BODY, size=10, bold=True, color=DARK)
        c.alignment = Alignment(vertical="center", wrap_text=True, indent=1)
        c.fill = PatternFill("solid", fgColor=SOLUTION)
        ws3.row_dimensions[r].height = 34
        r += 1
        c = ws3.cell(r, 2, url)
        c.font = Font(name=BODY, size=9, color=ACCENT, underline="single")
        c.hyperlink = url
        c.alignment = Alignment(vertical="center", indent=1)
        r += 1

        img = XLImage(path)
        ratio = target_w / float(img.width)
        img.width = target_w
        img.height = int(img.height * ratio)
        ws3.add_image(img, f"B{r}")
        rows_needed = int(img.height / 19) + 2
        for k in range(rows_needed):
            ws3.row_dimensions[r + k].height = 19
        r += rows_needed + 1

    # ============================================================ SHEET 4
    ws4 = wb.create_sheet("Sources")
    ws4.sheet_view.showGridLines = False
    for col, w in zip("ABC", (16, 44, 92)):
        ws4.column_dimensions[col].width = w

    r = 1
    ws4.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    c = ws4.cell(r, 1, "SOURCES — every URL used in this battlecard")
    c.font = Font(name=HEAD, size=16, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(vertical="center", indent=1)
    ws4.row_dimensions[r].height = 34
    r += 2

    for i, h in enumerate(["Vendor", "Key", "URL"], 1):
        c = ws4.cell(r, i, h)
        c.font = Font(name=HEAD, size=10, bold=True, color=WHITE)
        c.fill = PatternFill("solid", fgColor=DEEP)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BOX
    r += 1

    for vendor, table in (("AccuKnox", AK), ("Promptfoo", PF)):
        for key, url in table.items():
            ws4.cell(r, 1, vendor).font = Font(name=BODY, size=10, bold=True,
                                               color=NAVY)
            ws4.cell(r, 2, key).font = Font(name=BODY, size=10, color=DARK)
            c = ws4.cell(r, 3, url)
            c.font = Font(name=BODY, size=9, color=ACCENT, underline="single")
            c.hyperlink = url
            for col in (1, 2, 3):
                ws4.cell(r, col).border = BOX
                ws4.cell(r, col).alignment = Alignment(vertical="center",
                                                       indent=1)
            r += 1

    r += 1
    ws4.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    c = ws4.cell(r, 1,
                 "Promptfoo pages read 15 August 2026 with Firecrawl. A 'No' "
                 "for Promptfoo means the capability appears nowhere in their "
                 "site map of 280 URLs, nowhere in the six product pages they "
                 "list on their homepage, and nowhere in the red team, MCP, "
                 "model security, guardrails, code scanning, self-hosting, or "
                 "pricing documentation. Re-verify before any customer meeting.")
    c.font = Font(name=BODY, size=9, italic=True, color=MID)
    c.alignment = Alignment(vertical="center", wrap_text=True, indent=1)
    ws4.row_dimensions[r].height = 46

    # print setup so the sheet exports and prints cleanly
    for sheet in (ws, ws2, ws3, ws4):
        sheet.page_setup.orientation = "landscape"
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 0
        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.print_options.horizontalCentered = True
    ws.print_title_rows = "6:6"

    out = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                       "accuknox-vs-promptfoo-ai-security.xlsx")
    wb.save(out)
    print("saved:", out)
    print("rows:", len(ROWS))
    print("AK yes:", ak_yes, "PF yes:", pf_yes, "both:", both,
          "AK only:", ak_only, "PF only:", pf_only)


if __name__ == "__main__":
    build()
